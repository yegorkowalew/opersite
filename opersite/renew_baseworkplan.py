# -*- coding: utf-8 -*-
""" Работа с файлом План производства """
import django
import logging
from django.conf import settings
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "opersite.settings")
django.setup()

from baseworkplan.models import BaseWorkplan

file_path = settings.PRODUCTION_PLAN_FILE

col_names = ['ID', 'цех']
col_names_index = []

periods_list = ['Ц', 'М', 'К', 'З']
series_list = ['О', 'Z']

class ReadFindError(Exception):
   pass

def get_dates_list(ws, row_num, col_num):
    row_num = row_num-1
    dates_list = []
    for row in ws.iter_rows(min_row=row_num, min_col=col_num, max_row=row_num, max_col=ws.max_column):
        for cell in row:
            date_str = cell.value
            if isinstance(date_str, datetime):
                dates_list.append(date_str)
            else:
                if not dates_list:
                    raise ReadFindError('Что-то не так со строкой дат, значение: "%s" (строка: %s, столбец:%s)' % (date_str, row_num, cell))
                print('Что-то не так с датами, последняя %s, (строка: %s, столбец:%s)' % (dates_list[-1], row_num, cell))
                return dates_list
    return dates_list

def find_day(ws):
    # находим столбец с первым днем
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']:
                return [cell.row, cell.column]
    raise ReadFindError('Не нашел столбец "%s"' % day_list)

def find_names_col_num(ws, row_num, col_num, find_text):
    # находим номера нужных столбцов
    for col in range(1, col_num):
        if find_text == ws.cell(row=row_num, column=col).value:
            return col-1
    raise ReadFindError('Не нашел столбец "%s"' % find_text)

def get_word_dates(col_names_index, row, dates_list, periods_list, first_col):
    shop_dates = defaultdict(list)
    for i in range(first_col, len(row)):
        val = row[i-1].value
        if val:
            shop_dates[val].append(dates_list[i-first_col])
    
    for key, value in shop_dates.items():
        if key in periods_list:
            shop_dates[key] = [value[0], value[-1]]
    if not shop_dates:
        pass
    else:
        shop_dates['in_id'] = row[col_names_index[0]].value
        shop_dates['shop'] = row[col_names_index[1]].value
        return shop_dates

def to_dict(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb['График']
    try:
        first_day_coords = find_day(ws)
        for col_name in col_names:
            col_names_index.append(find_names_col_num(ws, first_day_coords[0], first_day_coords[1], col_name))
        dates_list = get_dates_list(ws, first_day_coords[0], first_day_coords[1])
    except ReadFindError as ind:
        print('Ошибка: %s' % ind)
        exit(0)

    full_dates = []
    
    for row in ws.iter_rows(min_row=first_day_coords[0]+1, max_col=ws.max_column+1, max_row=ws.max_row+1):
        zz = get_word_dates(col_names_index, row, dates_list, periods_list, first_day_coords[1])
        if zz:
            full_dates.append(zz)

    periods_list_list = []

    for i in full_dates:
        for shopitm in i.keys():
            if shopitm in periods_list:
                periods_list_list.append({'in_id':i['in_id'], 'letter':shopitm, 'shop':i['shop'], 'start_date':i[shopitm][0], 'end_date':i[shopitm][-1]})
            elif shopitm in series_list:
                for d_itm in i[shopitm]:
                    periods_list_list.append({'in_id':i['in_id'], 'letter':shopitm, 'shop':i['shop'], 'start_date':d_itm, 'end_date':d_itm})

    mdf = pd.DataFrame(periods_list_list)
    df_records = mdf.to_dict('records')
    return df_records

def insertToDB(df_records, m_obj):
    model_instances = [m_obj(**record) for record in df_records]
    try:
        m_obj.objects.bulk_create(model_instances)
    except Exception as identifier:
        print("insertToDB error with class: %s, %s" % (m_obj.__class__.__name__, identifier))

def renew_baseworkplan_worker():
    df = to_dict(file_path)
    BaseWorkplan.objects.all().delete()
    insertToDB(df, BaseWorkplan)

if __name__ == "__main__":
    renew_baseworkplan_worker()
    # df = to_dict(file_path)
    # from baseworkplan.models import BaseWorkplan
    # BaseWorkplan.objects.all().delete()
    # insertToDB(df, BaseWorkplan)
    # работает